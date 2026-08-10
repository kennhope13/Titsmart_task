import openpyxl
import shutil
import os

src_path = r"D:\HỆ THỐNG QUẢN LÝ CÔNG VIỆC-web-app-titsmart\documents\QL KH-VT; CHI PHÍ DỰ ÁN ĐẮK R'LẤP.xlsx"
dest_path = r"D:\HỆ THỐNG QUẢN LÝ CÔNG VIỆC-web-app-titsmart\documents\QL KH-VT; CHI PHÍ DỰ ÁN ĐẮK R'LẤP_SPLIT.xlsx"

wb = openpyxl.load_workbook(src_path)

def copy_and_rename(original_name, new_name):
    source = wb[original_name]
    copied = wb.copy_worksheet(source)
    copied.title = new_name
    return copied

# 1. Split THEO DÕI KẾ HOẠCH VẬT TƯ
kh_tech = copy_and_rename('THEO DÕI KẾ HOẠCH VẬT TƯ', 'KHVT - Kỹ thuật')
kh_progress = copy_and_rename('THEO DÕI KẾ HOẠCH VẬT TƯ', 'KHVT - Tiến độ đặt hàng')
kh_docs = copy_and_rename('THEO DÕI KẾ HOẠCH VẬT TƯ', 'KHVT - Chứng từ giao nhận')

# Clean 'KHVT - Kỹ thuật'
kh_tech.delete_cols(8, 11)
if kh_tech.max_column > 8:
    kh_tech.delete_cols(9, kh_tech.max_column - 8)

# Clean 'KHVT - Tiến độ đặt hàng'
kh_progress.delete_cols(14, 5)
kh_progress.delete_cols(5, 3)
if kh_progress.max_column > 11:
    kh_progress.delete_cols(12, kh_progress.max_column - 11)

# Clean 'KHVT - Chứng từ giao nhận'
kh_docs.delete_cols(5, 9)
if kh_docs.max_column > 10:
    kh_docs.delete_cols(11, kh_docs.max_column - 10)


# 2. Split THEO DÕI MUA SẮM HÀNG HÓA
ms_contract = copy_and_rename('THEO DÕI MUA SẮM HÀNG HÓA', 'Mua sắm - Hợp đồng')
ms_payment = copy_and_rename('THEO DÕI MUA SẮM HÀNG HÓA', 'Mua sắm - Thanh toán')

# Clean 'Mua sắm - Hợp đồng'
ms_contract.delete_cols(14, 2)
ms_contract.delete_cols(10, 2)
if ms_contract.max_column > 12:
    ms_contract.delete_cols(13, ms_contract.max_column - 12)

# Clean 'Mua sắm - Thanh toán'
ms_payment.delete_cols(4, 5)
if ms_payment.max_column > 11:
    ms_payment.delete_cols(12, ms_payment.max_column - 11)


# Remove original wide sheets
wb.remove(wb['THEO DÕI KẾ HOẠCH VẬT TƯ'])
wb.remove(wb['THEO DÕI MUA SẮM HÀNG HÓA'])

new_order = [
    'TỔNG QUAN',
    'KHVT - Kỹ thuật',
    'KHVT - Tiến độ đặt hàng',
    'KHVT - Chứng từ giao nhận',
    'Mua sắm - Hợp đồng',
    'Mua sắm - Thanh toán',
    'THEO DÕI CHI PHÍ CT',
    'TT Công Nhật',
    'Theo dõi chứng từ'
]

sheet_dict = {s.title: s for s in wb._sheets}
wb._sheets = [sheet_dict[name] for name in new_order if name in sheet_dict]

wb.save(dest_path)
wb.close()
print("Success")
